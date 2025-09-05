from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
from difflib import get_close_matches

# 9/4/25 note: Prior to this, I used the statement below to find the total extracted from ended TIFs
# df.sort_values("tif_year").groupby("tif_name").tail(1).query("end_year <= 2023")["cumulative_property_tax_extraction"].sum()

# ---------------------------
# FUNCTIONS
# ---------------------------
def urlList(url, year):
    """Return dict mapping TIF name → PDF URL using BeautifulSoup, ensuring no duplicates."""
    yr = str(year)[-2:]
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    tif_dict = {}
    seen_urls = set()
    for link in soup.find_all('a', href=lambda href: href and href.endswith(f'AR{yr}.pdf')):
        tif_name = link.get_text(strip=True)
        tif_url = "https://www.chicago.gov" + link['href']
        if tif_url not in seen_urls:
            tif_dict[tif_name] = tif_url
            seen_urls.add(tif_url)

    # Remove invalid "Archer Courts" URL if present
    archerCourtsUrlToRemove = f'https://www.chicago.gov/content/dam/city/depts/dcd/tif/{yr}reports/T_067_ArcherCourtsAR{yr}.pdf'
    if archerCourtsUrlToRemove in seen_urls and int(year) > 2022:
        tif_dict = {k: v for k, v in tif_dict.items() if v != archerCourtsUrlToRemove}

    return tif_dict

# ---------------------------
# CONFIG
# ---------------------------
url = 'https://www.chicago.gov/city/en/depts/dcd/supp_info/district-annual-reports--2024-.html'
year = 2024
excel_path = r"C:\Users\w\clonedGitRepos\chi-tif-parser\csvs\Chicago_2024_TIF_Illumination.xlsx"
sheet_name = "Chicago_2024_TIF_Illumination"

# ---------------------------
# LOAD URLS
# ---------------------------
urlList = urlList(url, year)

# ---------------------------
# LOAD EXCEL & TABLE
# ---------------------------
wb = load_workbook(excel_path)
ws = wb[sheet_name]

tif_table = None
for tbl in ws._tables.values():
    if tbl.name == "TIFs":
        tif_table = tbl
        break
if tif_table is None:
    raise ValueError("TIFs table not found.")

# Table boundaries
start_cell, end_cell = tif_table.ref.split(":")
start_row = int(''.join(filter(str.isdigit, start_cell)))
end_row = int(''.join(filter(str.isdigit, end_cell)))
start_col_letter = ''.join(filter(str.isalpha, start_cell))
end_col_letter = ''.join(filter(str.isalpha, end_cell))

# Convert column letters to numbers
from openpyxl.utils import column_index_from_string
start_col = column_index_from_string(start_col_letter)
end_col = column_index_from_string(end_col_letter)

# Find "TIF Name" column
tif_name_col = None
for col in range(start_col, end_col + 1):
    if ws.cell(row=start_row, column=col).value == "TIF Name":
        tif_name_col = col
        break
if tif_name_col is None:
    raise ValueError("'TIF Name' column not found in header row.")

# ---------------------------
# TRACKING
# ---------------------------
successes = []
fuzzy_matches = {}
substring_matches = {}
remaining_failures = []
used_keys = set()

# ---------------------------
# ITERATE TABLE ROWS
# ---------------------------
for row in range(start_row + 1, end_row + 1):
    cell = ws.cell(row=row, column=tif_name_col)
    tif_name = cell.value
    if not isinstance(tif_name, str):
        continue

    matched_url = None
    matched_key = None

    # Exact substring match
    for key, url in urlList.items():
        if key in tif_name:
            matched_url = url
            matched_key = key
            substring_matches[tif_name] = matched_key  # Track substring match
            break

    # Fuzzy match
    if not matched_url:
        close_keys = get_close_matches(tif_name, urlList.keys(), n=1, cutoff=0.8)
        if close_keys:
            matched_key = close_keys[0]
            matched_url = urlList[matched_key]
            fuzzy_matches[tif_name] = matched_key

    if matched_url:
        cell.hyperlink = matched_url
        cell.style = "Hyperlink"
        successes.append(tif_name)
        used_keys.add(matched_key)
    else:
        remaining_failures.append(tif_name)

# ---------------------------
# REPORT
# ---------------------------
print(f"Total table rows: {end_row - start_row}")
print(f"Hyperlinks applied: {len(successes)}")
print(f"TIF names with no URL match: {len(remaining_failures)}")

print("\n=== Substring Matches Applied ===")
for sheet_name in sorted(substring_matches.keys()):
    print(f"{sheet_name} → {substring_matches[sheet_name]}")

print("\n=== Fuzzy Matches Applied ===")
for sheet_name in sorted(fuzzy_matches.keys()):
    print(f"{sheet_name} → {fuzzy_matches[sheet_name]}")

print("\n=== Remaining Unmatched TIF Names ===")
for sheet_name in sorted(remaining_failures):
    # Attempt to show the closest URL key if it exists
    closest = get_close_matches(sheet_name, urlList.keys(), n=1, cutoff=0.5)
    closest_str = closest[0] if closest else "No close match"
    print(f"{sheet_name} → {closest_str}")

print("\n=== URL List Keys Not Used ===")
unused_keys = set(urlList.keys()) - used_keys
for key in sorted(unused_keys):
    print(key)

# ---------------------------
# SAVE
# ---------------------------
wb.save(excel_path)
print("\nExcel file updated with hyperlinks.")
